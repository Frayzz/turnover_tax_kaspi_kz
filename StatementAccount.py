from openpyxl import load_workbook
from numbers import Number


def main():
    wb = load_workbook(
        'Выписка_по_счету.xlsx')
    sheets = wb.sheetnames

    for sheet_name in sheets:
        print(sheet_name)

    sheet = wb.active

    debet_sum, return_sum, index = process_sheet(sheet)

    deduction_returns = debet_sum - return_sum
    three_percent_tax = deduction_returns * 0.03

    print(f'Итого оборота: {format_number(debet_sum)}тг')
    print(f'Возвраты на сумму: {format_number(return_sum)}тг')
    print(
        f'Обороты с вычтенным возвратом: {format_number(deduction_returns)}тг')
    print(f'Чистые 3%: {format_number(three_percent_tax)}тг')


def process_sheet(sheet):
    debet_sum = 0
    return_sum = 0

    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row[1] == 'Итого обороты в валюте счета':
            break

        if index >= 14:
            if isinstance(row[2], Number):
                if 'Возврат' in row[8]:
                    return_sum += row[2]

            if isinstance(row[3], Number):
                if 'Возврат' not in row[8]:
                    debet_sum += row[3]

    return debet_sum, return_sum, index


def format_number(number):
    return '{:,}'.format(number).replace(',', ' ')


if __name__ == "__main__":
    main()

